
import openpyxl
import datetime

def update_pick_and_roll(excel_path, data_dict):
    """
    Actualiza el archivo Excel de rotulación con nuevos datos.
    data_dict: { 'seccion_nombre': [ (partido, fecha), ... ] o [ (equipo, pj, pg, pp, pts), ... ] }
    """
    try:
        wb = openpyxl.load_workbook(excel_path)
        
        for section_name, rows_data in data_dict.items():
            found = False
            # Ajustamos nombres de sección para mayor flexibilidad
            possible_titles = [section_name, f"CLASIFICACION {section_name}", f"PRÓXIMA JORNADA {section_name}"]
            
            for sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value in possible_titles:
                            # Los datos empiezan en la fila siguiente
                            start_row = cell.row + 1
                            # Si es clasificación, saltamos 1 más para el encabezado de columnas si existe
                            if "CLASIFICACION" in sheetname or "CLASIFICACION" in (cell.value or ""):
                                # Comprobamos si la siguiente fila tiene encabezados como "PJ"
                                if sheet.cell(row=start_row, column=cell.column + 1).value in ["PJ", "PARTIDOS"]:
                                    start_row += 1
                            
                            print(f"-> Actualizando {section_name} en {sheetname}, fila {start_row}")
                            
                            # Limpieza de seguridad (limpia 20 filas)
                            for r in range(start_row, start_row + 20):
                                for c in range(cell.column, cell.column + 5):
                                    sheet.cell(row=r, column=c).value = None
                            
                            # Inserción
                            for i, item_data in enumerate(rows_data):
                                for j, val in enumerate(item_data):
                                    sheet.cell(row=start_row + i, column=cell.column + j).value = val
                            
                            found = True
                            break
                    if found: break
                if found: break
            
            if not found:
                print(f"X No se encontró la sección: {section_name}")

        wb.save(excel_path)
        print("\n¡Excel 'DATOS PICK AND ROLL' actualizado correctamente!")
        
    except Exception as e:
        print(f"Error crítico al actualizar el Excel: {e}")

if __name__ == "__main__":
    # Este bloque es para pruebas o ejecución directa
    PATH_EXCEL = r'\\172.16.80.129\g\EXCEL PROGRAMAS\PICK AND ROLL\DATOS PICK AND ROLL.xlsx'
    # Ejemplo de estructura de datos
    # data = { "N1 ANDALUZA": [("EQUIPO A - EQUIPO B", "01/01/24 18:00")], ... }
    print("Módulo de actualización Pick and Roll listo.")
