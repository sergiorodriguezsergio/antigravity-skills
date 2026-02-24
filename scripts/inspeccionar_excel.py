
import openpyxl

def inspect_excel_layout(excel_path):
    print(f"Inspeccionando estructura de: {excel_path}\n")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Analizamos las hojas principales
        target_sheets = ['RESULTADOS JORNADA', 'CLASIFICACION']
        
        for sheetname in target_sheets:
            if sheetname in wb.sheetnames:
                sheet = wb[sheetname]
                print(f"--- Hoja: {sheetname} ---")
                # Leemos las primeras 30 filas y 10 columnas para entender el layout
                for r in range(1, 31):
                    row_data = []
                    for c in range(1, 11):
                        val = sheet.cell(row=r, column=c).value
                        row_data.append(str(val) if val is not None else "")
                    if any(row_data):
                        print(f"Fila {r:2}: {' | '.join(row_data)}")
                print("\n")
            else:
                print(f"⚠️ Hoja '{sheetname}' no encontrada.\n")
                
        wb.close()
    except Exception as e:
        print(f"❌ Error: {e}")

if __name__ == "__main__":
    PATH_EXCEL = r'\\172.16.80.129\g\EXCEL PROGRAMAS\PICK AND ROLL\DATOS PICK AND ROLL.xlsx'
    inspect_excel_layout(PATH_EXCEL)
