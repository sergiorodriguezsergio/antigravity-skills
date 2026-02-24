
import openpyxl

def refine_update_results(excel_path, data_n1, data_tercera, clas_n1, clas_tercera):
    print(f"Refinando actualización en: {excel_path}")
    try:
        wb = openpyxl.load_workbook(excel_path)
        
        # 1. ACTUALIZAR RESULTADOS JORNADA
        if 'RESULTADOS JORNADA' in wb.sheetnames:
            ws_res = wb['RESULTADOS JORNADA']
            
            # --- N1 ANDALUZA (Filas 2 a 8, Columnas A y B) ---
            print("Actualizando Resultados N1 (Filas 2-8)...")
            for i, (partido, resultado, fecha) in enumerate(data_n1[:7]):
                ws_res.cell(row=2 + i, column=1).value = partido
                ws_res.cell(row=2 + i, column=2).value = resultado
            
            # --- TERCERA FEB (Filas 13 a 19, Columnas A y B) ---
            print("Actualizando Resultados Tercera FEB (Filas 13-19)...")
            for i, (partido, resultado, fecha) in enumerate(data_tercera[:7]):
                ws_res.cell(row=13 + i, column=1).value = partido
                ws_res.cell(row=13 + i, column=2).value = resultado

        # 2. ACTUALIZAR CLASIFICACION
        if 'CLASIFICACION' in wb.sheetnames:
            ws_clas = wb['CLASIFICACION']
            
            # --- CLASIFICACION N1 (Filas 2 a 15, Columnas A a E) ---
            print("Actualizando Clasificación N1 (Filas 2-15)...")
            for i, row_data in enumerate(clas_n1[:14]):
                # row_data: (pos, equipo, pj, pg, pp, pts)
                for j, val in enumerate(row_data):
                    ws_clas.cell(row=2 + i, column=1 + j).value = val
            
            # --- CLASIFICACION TERCERA FEB ---
            # Nota: Según la inspección, la Tercera FEB no aparecía claramente 
            # justo debajo. Vamos a buscar el título "CLASIFICACION TERCERA FEB" 
            # o ponerla en un lugar seguro (ej: Fila 20 en adelante o donde corresponda)
            # Si no hay hueco claro, la pondremos a partir de la 18 por si acaso.
            print("Buscando hueco para Clasificación Tercera FEB...")
            start_tercera = 18 # Asumimos un salto tras N1
            # Pero en la inspección N1 llega a la 15. Dejamos una libre.
            start_row_tercera = 17
            ws_clas.cell(row=start_row_tercera, column=1).value = "CLASIFICACION TERCERA FEB"
            for i, row_data in enumerate(clas_tercera[:14]):
                for j, val in enumerate(row_data):
                    ws_clas.cell(row=start_row_tercera + 1 + i, column=1 + j).value = val

        wb.save(excel_path)
        print("\n✅ Excel ajustado con precisión de filas y columnas.")
        
    except Exception as e:
        print(f"❌ Error al refinar: {e}")

if __name__ == "__main__":
    PATH_EXCEL = r'c:\Users\Equipo1\Documents\Antigravity\Skills\scripts\temp_pick_and_roll.xlsx'
    
    # Datos Jornada 19
    n1_res = [
        ("CB GIBRALEÓN - CB CIUDAD DE PALOS", "80-60", ""),
        ("ATICA SEVILLA CB CORIA - CB LEPE ALIUS", "89-60", ""),
        ("BARNETO MODAS LA PALMA 95 - NAUTICO SEVILLA", "74-56", ""),
        ("CD BALONCESTO HUELVA LA LUZ - AVATEL ÉCIJA", "67-100", ""),
        ("CB FRESAS - XEREZ CLUB DEPORTIVO", "59-64", ""),
        ("CDBC PILAS - CIRCULO MERCANTIL", "57-86", ""),
        ("RC LABRADORES - AD REM ONUBA", "APLAZADO", "")
    ]
    
    tercera_res = [
        ("CBA SPAIN - DEHESAS R. PEÑARROYA", "54-87", ""),
        ("HACHE P. MORALEJA - BAUBLOCK GYMNÁSTICA", "58-81", ""),
        ("INSOLAC CB ALCALÁ - BOSCO MÉRIDA PH", "50-83", ""),
        ("SAN ANTONIO CÁCERES - CB DOS HERMANAS", "62-73", ""),
        ("HUELVA COMERCIO VIRIDIS - BC BADAJOZ", "52-73", ""),
        ("CB SAN FERNANDO - LITHIUM I. SAGRADO", "85-79", ""),
        ("ATICA SEVILLA CB CORIA - S.D. ALJARAQUE", "APLAZADO", "")
    ]
    
    n1_clas = [
        ("1", "AD REM ONUBA", "17", "16", "1", "33"),
        ("2", "CIRCULO MERCANTIL", "18", "15", "3", "33"),
        ("3", "ATICA SEVILLA CB CORIA", "18", "12", "6", "30"),
        ("4", "CB FRESAS - SAFA REYES", "19", "10", "9", "29"),
        ("5", "BARNETO MODAS LA PALMA 95", "17", "12", "5", "29"),
        ("6", "CLUB NAUTICO SEVILLA", "18", "10", "8", "28"),
        ("7", "CB LEPE ALIUS EL JAMÓN", "17", "10", "7", "27"),
        ("8", "XEREZ CLUB DEPORTIVO", "19", "8", "11", "27"),
        ("9", "CDBC PILAS", "17", "9", "8", "26"),
        ("10", "CB GIBRALEÓN", "18", "8", "10", "26"),
        ("11", "RC LABRADORES", "16", "8", "8", "24"),
        ("12", "ÉCIJA BASKET", "18", "5", "13", "23"),
        ("13", "CB CIUDAD DE PALOS", "20", "2", "18", "22"),
        ("14", "HUELVA LA LUZ", "18", "0", "18", "18")
    ]
    
    tercera_clas = [
        ("1", "VÍTALY LA MAR BC BADAJOZ", "18", "16", "2", "34"),
        ("2", "BOSCO MÉRIDA PH", "18", "14", "4", "32"),
        ("3", "BAUBLOCK GYMNÁSTICA", "19", "13", "6", "32"),
        ("4", "CB SAN FERNANDO", "19", "11", "8", "30"),
        ("5", "SAN ANTONIO CÁCERES", "18", "11", "7", "29"),
        ("6", "LITHIUM IBERIA SAGRADO", "18", "11", "7", "29"),
        ("7", "HUELVA COMERCIO VIRIDIS", "18", "10", "8", "28"),
        ("8", "ATICA SEVILLA CB CORIA", "17", "10", "7", "27"),
        ("9", "INSOLAC CB ALCALÁ CAJA87", "18", "8", "10", "26"),
        ("10", "D.R. PEÑARROYA", "18", "7", "11", "25"),
        ("11", "EIFFAGE CB DOS HERMANAS", "17", "6", "11", "23"),
        ("12", "CBA SPAIN", "18", "3", "15", "21"),
        ("13", "HACHE P. MORALEJA", "18", "2", "16", "20"),
        ("14", "S.D. ALJARAQUE", "16", "3", "13", "19")
    ]
    
    refine_update_results(PATH_EXCEL, n1_res, tercera_res, n1_clas, tercera_clas)
